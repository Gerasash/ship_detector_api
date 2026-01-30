from datetime import datetime
from io import BytesIO
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    def __init__(self, results: dict, filename: str = "report"):
        self.results = results
        self.filename = filename
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


    def generate_pdf(self) -> BytesIO:
        """Генерирует простой PDF отчет"""
        logger.info("Starting PDF generation...")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20, bottomMargin=20)
        story = []
        styles = getSampleStyleSheet()

        # Заголовок (на английском)
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Normal'],
            fontSize=20,
            textColor=colors.black,
            spaceAfter=12,
            alignment=1,
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph("SHIP DETECTION REPORT", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Простой текст (на английском)
        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=6,
            fontName='Helvetica'
        )
        
        story.append(Paragraph(f"<b>Date/Time:</b> {self.timestamp}", body_style))
        story.append(Paragraph(f"<b>Processing Time:</b> {self.results.get('processing_time', 0):.3f}s", body_style))
        story.append(Paragraph(f"<b>Ships Detected:</b> {self.results.get('total_ships', 0)}", body_style))
        
        story.append(Spacer(1, 0.3 * inch))
        
        # Таблица детекций БЕЗ кириллицы
        if self.results.get('ships') and len(self.results.get('ships', [])) > 0:
            story.append(Paragraph("<b>DETECTIONS:</b>", body_style))
            story.append(Spacer(1, 0.1 * inch))
            
            data = [["ID", "Class", "Confidence (%)", "Coordinates"]]
            
            for i, ship in enumerate(self.results.get('ships', [])[:15], 1):
                x1, y1, x2, y2 = ship.get('bbox', [0, 0, 0, 0])
                conf_pct = f"{ship.get('conf', 0)*100:.1f}"
                coords = f"({int(x1)},{int(y1)},{int(x2)},{int(y2)})"
                
                data.append([
                    str(i),
                    ship.get('class', 'unknown'),
                    conf_pct,
                    coords
                ])
            
            table = Table(data, colWidths=[0.5*inch, 1.5*inch, 1.5*inch, 1.8*inch])
            
            # МИНИМАЛЬНЫЕ стили
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LINEBEFORE', (0, 0), (0, -1), 0.5, colors.grey),
                ('LINEAFTER', (-1, 0), (-1, -1), 0.5, colors.grey),
                ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
                ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)

        logger.info("Building PDF...")
        doc.build(story)
        buffer.seek(0)
        logger.info("PDF generated successfully")
        
        return buffer


    
    def generate_excel(self) -> BytesIO:
        """Генерирует Excel отчет с результатами детекции"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты"

        # Стили
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок
        ws['A1'] = "Отчет о детекции судов"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = center_alignment

        # Информация об обработке
        ws['A3'] = "Параметр"
        ws['B3'] = "Значение"
        ws['A3'].fill = header_fill
        ws['B3'].fill = header_fill
        ws['A3'].font = header_font
        ws['B3'].font = header_font

        info_rows = [
            ("Дата/время", self.timestamp),
            ("Время обработки", f"{self.results.get('processing_time', 0):.3f}с"),
            ("Обнаружено судов", str(self.results.get('total_ships', 0))),
        ]

        for idx, (param, value) in enumerate(info_rows, start=4):
            ws[f'A{idx}'] = param
            ws[f'B{idx}'] = value
            ws[f'A{idx}'].border = border
            ws[f'B{idx}'].border = border

        # Таблица с детекциями
        if self.results.get('ships'):
            start_row = len(info_rows) + 5

            headers = ["№", "Класс", "Уверенность (%)", "X1", "Y1", "X2", "Y2"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            for row_idx, ship in enumerate(self.results.get('ships', []), start=start_row + 1):
                x1, y1, x2, y2 = ship.get('bbox', [0, 0, 0, 0])
                row_data = [
                    row_idx - start_row,
                    ship.get('class', 'unknown'),
                    f"{ship.get('conf', 0)*100:.1f}",
                    int(x1),
                    int(y1),
                    int(x2),
                    int(y2)
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = center_alignment
                    cell.border = border

        # Автоширина колонок
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
